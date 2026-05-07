#!/usr/bin/env python3
"""Oracle Listener Log Analyzer — Web Application"""

import os, re, uuid, json, time, tempfile, threading, sqlite3, gzip, tarfile, shutil
from datetime import datetime
from flask import Flask, request, jsonify, render_template, Response, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500 MB (compressed upload limit)

SESSIONS: dict = {}
_LOCK = threading.Lock()

# ── Parsing ────────────────────────────────────────────────────────────────

MONTH_MAP = {m: f'{i+1:02d}' for i, m in enumerate(
    ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC'])}

DDL = """
CREATE TABLE IF NOT EXISTS connections (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp TEXT, date TEXT, hour INTEGER,
    service_name TEXT, instance_name TEXT, server_type TEXT,
    program TEXT, client_host TEXT, os_user TEXT,
    client_ip TEXT, client_port TEXT, action TEXT, status_code TEXT, success INTEGER
);
CREATE INDEX IF NOT EXISTS i1 ON connections(client_ip);
CREATE INDEX IF NOT EXISTS i2 ON connections(program);
CREATE INDEX IF NOT EXISTS i3 ON connections(client_host);
CREATE INDEX IF NOT EXISTS i4 ON connections(date);
CREATE INDEX IF NOT EXISTS i5 ON connections(service_name);
"""

_INSERT = (
    "INSERT INTO connections VALUES "
    "(NULL,:timestamp,:date,:hour,:service_name,:instance_name,:server_type,"
    ":program,:client_host,:os_user,:client_ip,:client_port,:action,:status_code,:success)"
)

# Pattern used to detect Oracle listener log header lines
LISTENER_HDR_RE = re.compile(
    r'^\d{2}-(?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)-\d{4}'
    r' \d{2}:\d{2}:\d{2}',
    re.IGNORECASE
)


def _ts(s):
    try:
        d, m, rest = s.split('-', 2)
        y, t = rest.split(' ', 1)
        return datetime.strptime(
            f"{y}-{MONTH_MAP[m.upper()]}-{int(d):02d} {t}", "%Y-%m-%d %H:%M:%S"
        )
    except Exception:
        return None


def _parse(line):
    line = line.strip()
    if not line or ' * ' not in line:
        return None
    p = line.split(' * ')
    if len(p) < 5:
        return None
    m = re.match(r'^(\d{2}-\w{3}-\d{4} \d{2}:\d{2}:\d{2})', p[0])
    if not m:
        return None
    ts = _ts(m.group(1))
    if not ts:
        return None

    cd, addr, act = p[1], p[2], p[3].strip().lower()
    if act not in ('establish', 'refused'):
        return None

    svc  = re.search(r'SERVICE_NAME=([^)]+)', cd)
    inst = re.search(r'INSTANCE_NAME=([^)]+)', cd)
    srv  = re.search(r'(?:^|\()SERVER=([A-Za-z]+)', cd)

    prog = host = user = None
    pos = cd.find('(CID=')
    if pos >= 0:
        c = cd[pos + 5:]
        pm = re.search(r'\(PROGRAM=([^)]+)\)', c)
        hm = re.search(r'\(HOST=([^)]+)\)', c)
        um = re.search(r'\(USER=([^)]+)\)', c)
        prog = pm.group(1).strip() if pm else None
        host = hm.group(1).strip() if hm else None
        user = um.group(1).strip() if um else None
        if host == '__jdbc__':
            host = None

    ah = re.search(r'\(HOST=([^)]+)\)', addr)
    ap = re.search(r'\(PORT=(\d+)\)', addr)
    sc = p[-1].strip()
    if not re.fullmatch(r'-?\d+', sc):
        sc = None

    return dict(
        timestamp    = ts.isoformat(sep=' '),
        date         = ts.date().isoformat(),
        hour         = ts.hour,
        service_name = svc.group(1).strip()  if svc  else None,
        instance_name= inst.group(1).strip() if inst else None,
        server_type  = srv.group(1).upper()  if srv  else None,
        program      = prog,
        client_host  = host,
        os_user      = user,
        client_ip    = ah.group(1).strip() if ah else None,
        client_port  = ap.group(1)         if ap else None,
        action       = act,
        status_code  = sc,
        success      = 1 if sc == '0' else 0,
    )


def _summary(conn):
    q = lambda sql: (conn.execute(sql).fetchone() or [0])[0] or 0
    return dict(
        total      = q("SELECT COUNT(*) FROM connections"),
        success    = q("SELECT SUM(success) FROM connections"),
        failures   = q("SELECT COUNT(*)-SUM(success) FROM connections"),
        uniq_ips   = q("SELECT COUNT(DISTINCT client_ip)    FROM connections WHERE client_ip    IS NOT NULL"),
        uniq_prog  = q("SELECT COUNT(DISTINCT program)      FROM connections WHERE program      IS NOT NULL"),
        uniq_host  = q("SELECT COUNT(DISTINCT client_host)  FROM connections WHERE client_host  IS NOT NULL"),
        uniq_user  = q("SELECT COUNT(DISTINCT os_user)      FROM connections WHERE os_user      IS NOT NULL"),
        uniq_svc   = q("SELECT COUNT(DISTINCT service_name) FROM connections WHERE service_name IS NOT NULL"),
        first      = conn.execute("SELECT MIN(timestamp) FROM connections").fetchone()[0],
        last       = conn.execute("SELECT MAX(timestamp) FROM connections").fetchone()[0],
    )


# ── Compressed file handling ───────────────────────────────────────────────

def _looks_like_listener(fileobj, check_lines=40):
    """Return True if file-like object contains Oracle listener log lines."""
    matches = seen = 0
    for raw in fileobj:
        line = raw.decode('utf-8', errors='replace') if isinstance(raw, bytes) else raw
        stripped = line.strip()
        if stripped:
            seen += 1
            if LISTENER_HDR_RE.match(stripped) and ' * ' in stripped:
                matches += 1
        if seen >= check_lines:
            break
    return matches > 0


def _scan_tar(archive_path):
    """
    Open a .tar.gz and return [(member_name, uncompressed_size)] for
    qualifying listener log members. Does not extract to disk.
    """
    qualifying = []
    try:
        with tarfile.open(archive_path, 'r:gz') as tar:
            for member in tar.getmembers():
                if not member.isfile():
                    continue
                ext = os.path.splitext(member.name.lower())[1]
                if ext and ext not in ('.log', '.txt', ''):
                    continue
                fobj = tar.extractfile(member)
                if fobj is None:
                    continue
                ok = _looks_like_listener(fobj)
                fobj.close()
                if ok:
                    qualifying.append((member.name, member.size))
    except Exception:
        pass
    return qualifying


def _decompress_gz(src_path, src_name):
    """
    Streaming decompress a .gz file to a temp file (1 MB chunks).
    Returns the temp path, or None if the content is not a listener log.
    Deletes src_path on exit regardless of outcome.
    """
    fd, dst = tempfile.mkstemp(suffix='.log')
    os.close(fd)
    try:
        with gzip.open(src_path, 'rb') as gz, open(dst, 'wb') as out:
            shutil.copyfileobj(gz, out, length=1024 * 1024)
        with open(dst, 'r', encoding='utf-8', errors='replace') as f:
            if not _looks_like_listener(f):
                os.unlink(dst)
                return None
        return dst
    except Exception:
        try:
            os.unlink(dst)
        except Exception:
            pass
        return None
    finally:
        try:
            os.unlink(src_path)
        except Exception:
            pass


def _resolve_upload(upload_file):
    """
    Save one uploaded file to temp and prepare it for the worker.
    Returns a list of log sources:
      - str  → plain temp file path
      - ('tar', archive_path, [(name, size), ...]) → streaming tar source
    Returns [] if no listener logs were found.
    """
    name = upload_file.filename or 'upload'
    nl   = name.lower()

    fd, src = tempfile.mkstemp(suffix='.tmp')
    os.close(fd)
    upload_file.save(src)

    if nl.endswith('.tar.gz') or nl.endswith('.tgz'):
        members = _scan_tar(src)
        if not members:
            try:
                os.unlink(src)
            except Exception:
                pass
            return []
        return [('tar', src, members)]

    if nl.endswith('.gz'):
        result = _decompress_gz(src, name)  # deletes src internally
        return [result] if result else []

    # Plain .log / .txt — verify it's a listener log before accepting
    with open(src, 'r', encoding='utf-8', errors='replace') as f:
        if not _looks_like_listener(f):
            try:
                os.unlink(src)
            except Exception:
                pass
            return []
    return [src]


# ── Fingerprinting & compatibility ─────────────────────────────────────────

def _instance_base(name):
    """Strip trailing digits from instance name: ORCL1 → ORCL."""
    return name.rstrip('0123456789') or name


def _fingerprint_source(src, max_records=2000):
    """
    Scan up to max_records parsed lines from a log source and return
    {'services': set, 'instances': set}.
    Works with both plain file paths and tar sources.
    """
    services, instances = set(), set()
    count = [0]

    def scan_file(path):
        with open(path, encoding='utf-8', errors='replace') as f:
            for line in f:
                r = _parse(line)
                if r:
                    if r['service_name']:  services.add(r['service_name'].upper())
                    if r['instance_name']: instances.add(r['instance_name'].upper())
                    count[0] += 1
                    if count[0] >= max_records:
                        return

    if isinstance(src, str):
        scan_file(src)
    else:
        _, archive_path, members = src
        try:
            with tarfile.open(archive_path, 'r:gz') as tar:
                member_set = {m for m, _ in members}
                for tarinfo in tar:
                    if count[0] >= max_records:
                        break
                    if tarinfo.name not in member_set or not tarinfo.isfile():
                        continue
                    fobj = tar.extractfile(tarinfo)
                    if fobj is None:
                        continue
                    fd, tp = tempfile.mkstemp(suffix='.log')
                    os.close(fd)
                    try:
                        with open(tp, 'wb') as out:
                            shutil.copyfileobj(fobj, out, 1024 * 1024)
                        fobj.close()
                        scan_file(tp)
                    finally:
                        try:
                            os.unlink(tp)
                        except Exception:
                            pass
        except Exception:
            pass

    return {'services': services, 'instances': instances}


def _compatible_databases(fps):
    """
    Return (True, None) if files appear to be from the same DB,
    or (False, reason_string) if they look like different databases.
    """
    if len(fps) < 2:
        return True, None

    common_svc = fps[0]['services'].copy()
    for fp in fps[1:]:
        common_svc &= fp['services']
    if common_svc:
        return True, None

    def bases(fp):
        return {_instance_base(i) for i in fp['instances']}

    all_bases = [bases(fp) for fp in fps]
    common_bases = all_bases[0].copy()
    for b in all_bases[1:]:
        common_bases &= b
    if common_bases:
        return True, None

    summaries = []
    for i, fp in enumerate(fps, 1):
        svc  = ', '.join(sorted(fp['services'])[:4])  or '(none)'
        inst = ', '.join(sorted(fp['instances'])[:4]) or '(none)'
        summaries.append(f'File {i}: services=[{svc}] instances=[{inst}]')
    reason = ('No common service names or instance name prefixes found.\n'
              + '\n'.join(summaries))
    return False, reason


# ── Worker ─────────────────────────────────────────────────────────────────

def _worker(sid, log_sources, db_path):
    """
    log_sources: list of:
      - str                                  → plain temp file path
      - ('tar', archive_path, [(name,size)]) → streaming tar extraction
    Files/archives are deleted after processing to free disk space.
    """
    try:
        conn = sqlite3.connect(db_path)
        conn.executescript(DDL)

        # Count total lines (plain files counted directly;
        # tar members estimated from uncompressed size ~150 bytes/line)
        total = 0
        for src in log_sources:
            if isinstance(src, str):
                with open(src, encoding='utf-8', errors='replace') as f:
                    total += sum(1 for _ in f)
            else:
                _, _, members = src
                total += max(sum(sz // 150 for _, sz in members), 1)
        total = max(total, 1)
        with _LOCK:
            SESSIONS[sid]['total_lines'] = total

        batch, parsed, done = [], 0, 0

        for src in log_sources:
            if isinstance(src, str):
                with open(src, encoding='utf-8', errors='replace') as f:
                    for raw in f:
                        done += 1
                        r = _parse(raw)
                        if r:
                            batch.append(r)
                            parsed += 1
                        if len(batch) >= 5000:
                            conn.executemany(_INSERT, batch)
                            conn.commit()
                            batch.clear()
                        if done % 10000 == 0:
                            with _LOCK:
                                SESSIONS[sid].update(
                                    parsed=parsed, lines_done=done,
                                    pct=min(int(100 * done / total), 99))
                try:
                    os.unlink(src)
                except Exception:
                    pass

            else:
                _, archive_path, members = src
                member_set = {m for m, _ in members}
                try:
                    with tarfile.open(archive_path, 'r:gz') as tar:
                        for tarinfo in tar:
                            if tarinfo.name not in member_set or not tarinfo.isfile():
                                continue
                            fobj = tar.extractfile(tarinfo)
                            if fobj is None:
                                continue
                            # Extract this member to a temp file, then process and delete
                            fd, tp = tempfile.mkstemp(suffix='.log')
                            os.close(fd)
                            with open(tp, 'wb') as out:
                                shutil.copyfileobj(fobj, out, 1024 * 1024)
                            fobj.close()
                            try:
                                with open(tp, encoding='utf-8', errors='replace') as f:
                                    for raw in f:
                                        done += 1
                                        r = _parse(raw)
                                        if r:
                                            batch.append(r)
                                            parsed += 1
                                        if len(batch) >= 5000:
                                            conn.executemany(_INSERT, batch)
                                            conn.commit()
                                            batch.clear()
                                        if done % 10000 == 0:
                                            with _LOCK:
                                                SESSIONS[sid].update(
                                                    parsed=parsed, lines_done=done,
                                                    pct=min(int(100 * done / total), 99))
                            finally:
                                try:
                                    os.unlink(tp)
                                except Exception:
                                    pass
                finally:
                    try:
                        os.unlink(archive_path)
                    except Exception:
                        pass

        if batch:
            conn.executemany(_INSERT, batch)
            conn.commit()

        sm = _summary(conn)
        conn.close()
        with _LOCK:
            SESSIONS[sid].update(status='ready', parsed=parsed,
                                 lines_done=done, pct=100, summary=sm)
    except Exception as e:
        with _LOCK:
            SESSIONS[sid].update(status='error', error=str(e))
    finally:
        # Safety cleanup for any sources not yet deleted
        for src in log_sources:
            if isinstance(src, str):
                try:
                    os.unlink(src)
                except Exception:
                    pass
            elif isinstance(src, tuple) and len(src) >= 2:
                try:
                    os.unlink(src[1])
                except Exception:
                    pass


# ── Query builder ──────────────────────────────────────────────────────────

DIMS = [
    ('client_ip',    'Client IP'),
    ('program',      'Program'),
    ('client_host',  'Client Host'),
    ('os_user',      'OS User'),
    ('service_name', 'Service'),
    ('instance_name','Instance'),
    ('server_type',  'Server Type'),
    ('date',         'Date'),
    ('hour',         'Hour'),
    ('action',       'Action'),
    ('status_code',  'Status Code'),
]
DIMKEYS   = [k for k, _ in DIMS]
COUNTABLE = ['client_ip', 'program', 'client_host', 'os_user', 'service_name']


def _build_sql(group_by, limit=2000):
    valid = [c for c in group_by if c in DIMKEYS] or ['client_ip']
    dlabel = dict(DIMS)
    sel, hdr = [], []

    for c in valid:
        sel.append(f"COALESCE(CAST({c} AS TEXT), '(unknown)') AS {c}")
        hdr.append(dlabel[c])

    sel += [
        "COUNT(*) AS connections",
        "SUM(success) AS success",
        "COUNT(*)-SUM(success) AS failures",
        "ROUND(100.0*SUM(success)/COUNT(*), 1) AS pct",
    ]
    hdr += ["Connections", "Success", "Failures", "% Success"]

    for c in COUNTABLE:
        if c not in valid:
            sel.append(f"COUNT(DISTINCT {c}) AS u_{c}")
            hdr.append(f"# {dlabel[c]}")

    sel += ["MIN(timestamp)", "MAX(timestamp)"]
    hdr += ["First Seen", "Last Seen"]

    gc  = [f"COALESCE(CAST({c} AS TEXT), '(unknown)')" for c in valid]
    sql = (f"SELECT {', '.join(sel)} FROM connections "
           f"GROUP BY {', '.join(gc)} ORDER BY connections DESC LIMIT {limit}")
    return sql, hdr


# ── Routes ─────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html', dims=DIMS)


@app.route('/upload', methods=['POST'])
def upload():
    files = request.files.getlist('file')
    files = [f for f in files if f and f.filename]
    if not files:
        return jsonify(error='No files provided'), 400

    all_sources, file_names, skipped = [], [], []
    for f in files:
        sources = _resolve_upload(f)
        if not sources:
            skipped.append(f.filename)
        else:
            all_sources.extend(sources)
            file_names.append(f.filename)

    if not all_sources:
        msg = ('No valid Oracle listener log files found in: '
               + ', '.join(skipped)) if skipped else 'No valid listener log files provided'
        return jsonify(error=msg), 400

    # Check compatibility — warn but do NOT hard-block
    warning_msg = None
    if len(all_sources) > 1:
        fps = [_fingerprint_source(s) for s in all_sources]
        ok, reason = _compatible_databases(fps)
        if not ok:
            warning_msg = reason

    sid = str(uuid.uuid4())
    fd, dp = tempfile.mkstemp(suffix='.db')
    os.close(fd)

    if warning_msg:
        with _LOCK:
            SESSIONS[sid] = dict(
                status='pending_confirm',
                log_sources=all_sources,
                db_path=dp,
                file_names=file_names,
                warning_msg=warning_msg,
            )
        return jsonify(warning=True, warning_msg=warning_msg, session_id=sid)

    with _LOCK:
        SESSIONS[sid] = dict(status='parsing', parsed=0, lines_done=0,
                             total_lines=1, pct=0, db_path=dp,
                             file_names=file_names)

    threading.Thread(target=_worker, args=(sid, all_sources, dp), daemon=True).start()
    return jsonify(session_id=sid, file_count=len(all_sources))


@app.route('/confirm/<sid>', methods=['POST'])
def confirm(sid):
    with _LOCK:
        s = SESSIONS.get(sid)
        if not s or s.get('status') != 'pending_confirm':
            return jsonify(error='Invalid session'), 400
        log_sources = s.pop('log_sources')
        s.pop('warning_msg', None)
        s.update(status='parsing', parsed=0, lines_done=0, total_lines=1, pct=0)

    threading.Thread(
        target=_worker,
        args=(sid, log_sources, s['db_path']),
        daemon=True
    ).start()
    return jsonify(ok=True)


@app.route('/cancel/<sid>', methods=['POST'])
def cancel_session(sid):
    with _LOCK:
        s = SESSIONS.pop(sid, None)
    if s:
        for src in s.get('log_sources', []):
            if isinstance(src, str):
                try: os.unlink(src)
                except Exception: pass
            elif isinstance(src, tuple) and len(src) >= 2:
                try: os.unlink(src[1])
                except Exception: pass
        try: os.unlink(s.get('db_path', ''))
        except Exception: pass
    return jsonify(ok=True)


@app.route('/progress/<sid>')
def progress(sid):
    def gen():
        for _ in range(1200):   # up to 10 minutes
            with _LOCK:
                s = {k: v for k, v in SESSIONS.get(sid, {}).items()
                     if k not in ('db_path', 'log_sources')}
            if not s:
                yield f"data: {json.dumps({'error': 'session not found'})}\n\n"
                break
            yield f"data: {json.dumps(s)}\n\n"
            if s.get('status') in ('ready', 'error'):
                break
            time.sleep(0.5)
    return Response(gen(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


@app.route('/query/<sid>')
def query(sid):
    with _LOCK:
        s = SESSIONS.get(sid)
    if not s or s['status'] != 'ready':
        return jsonify(error='Session not ready'), 400

    sql, hdr = _build_sql(
        request.args.getlist('g'),
        min(int(request.args.get('limit', 2000)), 10000),
    )
    conn = sqlite3.connect(s['db_path'])
    rows = conn.execute(sql).fetchall()
    conn.close()
    return jsonify(headers=hdr, rows=[list(r) for r in rows])


@app.route('/export/<sid>')
def export(sid):
    with _LOCK:
        s = SESSIONS.get(sid)
    if not s or s['status'] != 'ready':
        return jsonify(error='Session not ready'), 400

    sql, hdr = _build_sql(request.args.getlist('g'), 100000)
    conn = sqlite3.connect(s['db_path'])
    rows = conn.execute(sql).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Analysis'

    hfill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True)
    ws.append(hdr)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')

    for row in rows:
        ws.append(list(row))

    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, 50)
    ws.freeze_panes = 'A2'

    fd, tmp = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(tmp)

    label = '_'.join(request.args.getlist('g')) or 'analysis'
    return send_file(
        tmp, as_attachment=True,
        download_name=f'listener_{label}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/ipmap/<sid>')
def ipmap(sid):
    with _LOCK:
        s = SESSIONS.get(sid)
    if not s or s['status'] != 'ready':
        return jsonify(error='Not ready'), 400

    conn = sqlite3.connect(s['db_path'])
    total = conn.execute("SELECT COUNT(*) FROM connections").fetchone()[0] or 1

    rows = conn.execute("""
        SELECT
            c1.client_ip,
            COUNT(*)                                                        AS connections,
            COALESCE(SUM(c1.success), 0)                                   AS success,
            COUNT(*) - COALESCE(SUM(c1.success), 0)                       AS failures,
            (SELECT c2.client_host FROM connections c2
             WHERE c2.client_ip = c1.client_ip AND c2.client_host IS NOT NULL
             GROUP BY c2.client_host ORDER BY COUNT(*) DESC LIMIT 1)       AS primary_host,
            COUNT(DISTINCT c1.program)                                     AS programs,
            MIN(c1.timestamp)                                              AS first_seen,
            MAX(c1.timestamp)                                              AS last_seen
        FROM connections c1
        WHERE c1.client_ip IS NOT NULL
        GROUP BY c1.client_ip
        ORDER BY connections DESC
    """).fetchall()
    conn.close()

    return jsonify(total=total, ips=[{
        'ip': r[0], 'connections': r[1], 'success': r[2], 'failures': r[3],
        'host': r[4], 'programs': r[5], 'first': r[6], 'last': r[7],
        'pct': round(100 * r[1] / total, 2),
    } for r in rows])


@app.route('/detail/<sid>')
def detail(sid):
    with _LOCK:
        s = SESSIONS.get(sid)
    if not s or s['status'] != 'ready':
        return jsonify(error='Not ready'), 400

    ip = request.args.get('ip', '').strip()
    if not ip:
        return jsonify(error='No IP specified'), 400

    conn = sqlite3.connect(s['db_path'])

    def q(sql, p=()):
        return [list(r) for r in conn.execute(sql, p).fetchall()]

    basic = conn.execute("""
        SELECT COUNT(*), COALESCE(SUM(success),0),
               COUNT(*)-COALESCE(SUM(success),0),
               COUNT(DISTINCT program), COUNT(DISTINCT client_host),
               COUNT(DISTINCT os_user), COUNT(DISTINCT service_name),
               MIN(timestamp), MAX(timestamp)
        FROM connections WHERE client_ip = ?
    """, (ip,)).fetchone()

    total_all = conn.execute("SELECT COUNT(*) FROM connections").fetchone()[0] or 1

    result = dict(
        ip=ip,
        total=basic[0], success=basic[1], failures=basic[2],
        uniq_programs=basic[3], uniq_hosts=basic[4],
        uniq_users=basic[5], uniq_services=basic[6],
        first=basic[7], last=basic[8],
        pct=round(100 * basic[0] / total_all, 2),
        programs  = q("SELECT COALESCE(program,'(unknown)'),COUNT(*),COALESCE(SUM(success),0) FROM connections WHERE client_ip=? GROUP BY program ORDER BY COUNT(*) DESC", (ip,)),
        hosts     = q("SELECT COALESCE(client_host,'(no hostname)'),COUNT(*) FROM connections WHERE client_ip=? GROUP BY client_host ORDER BY COUNT(*) DESC", (ip,)),
        users     = q("SELECT COALESCE(os_user,'(unknown)'),COUNT(*) FROM connections WHERE client_ip=? GROUP BY os_user ORDER BY COUNT(*) DESC", (ip,)),
        services  = q("SELECT COALESCE(service_name,'(unknown)'),COUNT(*) FROM connections WHERE client_ip=? GROUP BY service_name ORDER BY COUNT(*) DESC", (ip,)),
        by_hour   = q("SELECT hour,COUNT(*) FROM connections WHERE client_ip=? GROUP BY hour ORDER BY hour", (ip,)),
        by_date   = q("SELECT date,COUNT(*) FROM connections WHERE client_ip=? GROUP BY date ORDER BY date", (ip,)),
        fail_codes= q("SELECT status_code,COUNT(*) FROM connections WHERE client_ip=? AND success=0 GROUP BY status_code ORDER BY COUNT(*) DESC", (ip,)),
        by_hour_by_date = q("SELECT date,hour,COUNT(*) FROM connections WHERE client_ip=? GROUP BY date,hour ORDER BY date,hour", (ip,)),
    )
    conn.close()
    return jsonify(result)


if __name__ == '__main__':
    app.run(debug=True, port=5000, threaded=True)
